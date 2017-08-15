import os
import openpyxl
from .define import STRING_EMPTY, excel_column, excel_coordinate
import re

class ExcelFiledNameDescript(object):
    def __init__(self, **kwargs):
        self.field_desc = None
        self.orignal_str = STRING_EMPTY
        self.name = STRING_EMPTY
        return super().__init__(**kwargs)

    def init(self, field_desc, name_str):
        self.field_desc = field_desc
        self.orignal_str = name_str
        self.name = name_str
        return True


class EnumFieldType:
    Min = 0
    Bool = 1
    Int = 2
    Long = 3
    Float = 4
    Base_Max = 5
    String = 6
    Vec = 7
    Map = 8
    VecVec = 9
    MapVec = 10
    Max = 11

    Base_Type_Strs = {
        "bool": Bool,
        "int": Int,
        "long": Long,
        "float": Float
        }

    String_Str = "string"

    @staticmethod
    def is_base_type(field_type):
        return field_type > EnumFieldType.Min and field_type < EnumFieldType.Base_Max

    @staticmethod
    def is_complex_type(field_type):
        return field_type > EnumFieldType.Base_Max and field_type < EnumFieldType.Max

    @staticmethod
    def is_collection_type(field_type):
        return EnumFieldType.is_complex_type(field_type) and EnumFieldType.String != field_type

    def is_vec_collection_type(field_type):
        if EnumFieldType.is_collection_type(field_type):
            return EnumFieldType.Vec == field_type or EnumFieldType.VecVec == field_type
    
    def is_map_collection_type(field_type):
        if EnumFieldType.is_collection_type(field_type):
            return EnumFieldType.Map == field_type or EnumFieldType.MapVec == field_type
    

    @staticmethod
    def parse_type(type_str):
        ret = False
        field_type =  key_type = val_type = EnumFieldType.Min
        type_str = type_str.strip()
        if not ret: # base 
            ret, field_type = EnumFieldType._parse_base_type(type_str)
        if not ret: # string
            if type_str == EnumFieldType.String_Str:
                ret = True
                field_type = EnumFieldType.String
        if not ret: # vecvec
            m_ret = re.match(r"^\[\[(.*)\]\]$", type_str) 
            if m_ret: 
                sub_ret, sub_type = EnumFieldType._parse_base_type(m_ret.group(1))
                if sub_ret and EnumFieldType.is_base_type(sub_type):
                    field_type = EnumFieldType.VecVec
                    val_type = sub_type
                    ret = True
        if not ret: # mapvec
            m_ret = re.match(r"^<(.*), \[(.*)\]>$", type_str) 
            if m_ret:
                parse_succ = False
                sub_ret, sub_type = EnumFieldType._parse_base_type(m_ret.group(1))
                if sub_ret and EnumFieldType.is_base_type(sub_type):
                    field_type = EnumFieldType.MapVec
                    key_type = sub_type
                    sub_ret, sub_type = EnumFieldType._parse_base_type(m_ret.group(2))
                    if sub_ret and EnumFieldType.is_base_type(sub_type):
                        val_type = sub_type
                        parse_succ = True
                        ret = True
                if not parse_succ:
                    field_type =  key_type = val_type = EnumFieldType.Min
        if not ret: # vec
            m_ret = re.match(r"^\[(.*)\]$", type_str) 
            if m_ret: 
                sub_ret, sub_type = EnumFieldType._parse_base_type(m_ret.group(1))
                if sub_ret and EnumFieldType.is_base_type(sub_type):
                    field_type = EnumFieldType.Vec
                    val_type = sub_type
                    ret = True
        if not ret: # map
            m_ret = re.match(r"^<(.*), (.*)>$", type_str) 
            if m_ret:
                parse_succ = False
                sub_ret, sub_type  = EnumFieldType._parse_base_type(m_ret.group(1))
                if sub_ret and EnumFieldType.is_base_type(sub_type):
                    field_type = EnumFieldType.Map
                    key_type = sub_type
                    sub_ret, sub_type = EnumFieldType._parse_base_type(m_ret.group(2))
                    if sub_ret and EnumFieldType.is_base_type(sub_type):
                        val_type = sub_type
                        parse_succ = True
                        ret = True
                if not parse_succ:
                    field_type =  key_type = val_type = EnumFieldType.Min
        return True, field_type, key_type, val_type

    @staticmethod
    def _parse_base_type(type_str):
        if type_str in EnumFieldType.Base_Type_Strs:
            return True, EnumFieldType.Base_Type_Strs[type_str]
        return False, None


class ExcelFieldTypeDescript(object):
    def __init__(self, **kwargs):
        self.field_desc = None
        self.orignal_str = STRING_EMPTY
        self.field_type = EnumFieldType.Min
        self.field_key_type = EnumFieldType.Min
        self.field_val_type = EnumFieldType.Min
        self.is_key = False
        self.is_group = False
        return super().__init__(**kwargs)

    def init(self, field_desc, type_str):
        self.field_desc = field_desc
        self.orignal_str = type_str
        return self.parse()

    def parse(self):
        segment_strs = self.orignal_str.strip().split(';')
        if len(segment_strs) <= 0 or not segment_strs[0]:
            return False
        if not self._parse_type(segment_strs[0]):
            return False
        if len(segment_strs) > 1:
            for segment_str in segment_strs[1:]:
                self._parse_key(segment_str)
                self._parse_group(segment_str)
        return True
    
    def _parse_type(self, type_str):
        ret, self.field_type, self.field_key_type, self.field_val_type = EnumFieldType.parse_type(type_str)
        return ret

    def _parse_key(self, segment_str):
        if segment_str.lower() == "key" \
        and not EnumFieldType.is_collection_type(self.field_type):
            self.is_key = True

    def _parse_group(self, segment_str):
        if segment_str.lower() == "group" \
            and not EnumFieldType.is_collection_type(self.field_type):
            self.is_group = True


class ExcelFieldDescript(object):
    def __init__(self, **kwargs):
        self.excel_desc = None
        self.column = ""
        self.name_desc = None
        self.type_desc = None
        return super().__init__(**kwargs)

    def init(self, excel_desc, column, name_str, type_str):
        self.column = column
        self.name_desc = ExcelFiledNameDescript()
        self.type_desc = ExcelFieldTypeDescript()
        all_ok = True
        all_ok = all_ok and self.name_desc.init(self, name_str)
        all_ok = all_ok and self.type_desc.init(self, type_str)
        return all_ok

class ExcelExtraFieldDescript(object):
    def __init__(self, **kwargs):
        self.field_type = []
        field_name = STRING_EMPTY
        return super().__init__(**kwargs)

class ExcelDescript(object):
    START_ROW_CELL = "A1"
    EXTRA_FIELD_ROW = 2

    def __init__(self, **kwargs):
        self.start_row = -1
        self.min_column = -1
        self.max_column = -2
        self.field_descs = []
        self.sheet_data = None
        self.extra_field_descs = []
        return super().__init__(**kwargs)

    def init(self, sheet_data):
        self.sheet_data = sheet_data
        self.start_row = int(sheet_data[ExcelDescript.START_ROW_CELL].value)
        if self.start_row <= 0: 
            return False
        all_ok = True
        for cell in self.sheet_data[ExcelDescript.EXTRA_FIELD_ROW:ExcelDescript.EXTRA_FIELD_ROW]:
            if not cell.value or not cell.value.strip(): 
                continue
            (extra_field_type, extra_field_name) = cell.value.split(':')[0:2]
            if not extra_field_type or not extra_field_name:
                all_ok = False
                break
            extra_field_desc = ExcelExtraFieldDescript()
            extra_field_desc.field_type = extra_field_type.strip('.').split('.')
            extra_field_desc.field_name = extra_field_name
            self.extra_field_descs.append(extra_field_desc)
        if all_ok:
            for name_cell, type_cell in self.sheet_data.iter_cols(min_row=self.start_row, max_row=self.start_row+1):
                if not name_cell.value or not type_cell.value:
                    if self.min_column < 0: continue
                    else: break
                if self.min_column < 0: 
                    self.min_column = name_cell.col_idx
                self.max_column = name_cell.col_idx
                field_desc = ExcelFieldDescript()
                if field_desc.init(self, name_cell.col_idx, name_cell.value, type_cell.value):
                    self.field_descs.append(field_desc)
                else: 
                    all_ok = False
                    break
        return all_ok and len(self.field_descs) > 0
            
    @staticmethod
    def load(file_path, sheet_name):
        excel_desc = None
        if os.path.exists(file_path) and os.path.isfile(file_path):
            xls_data = openpyxl.load_workbook(file_path)
            if xls_data and sheet_name in xls_data.sheetnames:
                excel_desc = ExcelDescript()
                if not excel_desc.init(xls_data[sheet_name]):
                    excel_desc = None
        return excel_desc